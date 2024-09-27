import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from streamlit import session_state as state
import os
from serpapi import GoogleSearch
import openai

# Set your API keys as environment variables or directly in the code (not recommended for security reasons)
SERPAPI_API_KEY = os.getenv("SERPAPI_API_KEY")
openai.api_key = os.getenv("OPENAI_API_KEY")

st.set_page_config(layout="wide", page_title="Backlink Relevancy Analyser")

# Constants
NOT_ACCEPTABLE = 20
HIGH_RELEVANCY = 40

# RS Online description
RS_ONLINE_DESCRIPTION = """
RS Online, or RS Components, is a global distributor of industrial and electronic products, serving a broad range of sectors and industries with a comprehensive selection of tools, equipment, and components.

Key Target Sectors: Manufacturing and Industrial, Engineering and R&D, Healthcare and Pharmaceuticals, Energy and Utilities, Construction and Facilities Management.

Product Ranges: Connectors, Electrification, Cables & Wires, Test & Measurement Instruments, Automation & Control, Safety, Mechanical & Fluid Power, Facilities & Maintenance, Semiconductors.

RS also offers technical support, procurement solutions, design services, and calibration services for industrial and technical needs.
"""

def fetch_domain_search_results(domain: str) -> list:
    params = {
        "engine": "google",
        "q": f"site:{domain}",
        "api_key": SERPAPI_API_KEY,
        "num": 10
    }
    search = GoogleSearch(params)
    try:
        results = search.get_dict()
        organic_results = results.get("organic_results", [])
        extracted_results = []
        for item in organic_results:
            extracted_results.append({
                'title': item.get('title', ''),
                'snippet': item.get('snippet', '')
            })
        return extracted_results
    except Exception as e:
        st.error(f"Error fetching search results for {domain}: {e}")
        return []

def generate_domain_explanation_from_results(results: list) -> str:
    if not results:
        return "No information available about this domain."
    # Combine titles and snippets to create a summary
    combined_text = ' '.join([f"{item['title']}. {item['snippet']}" for item in results])
    # Use OpenAI API to summarize
    prompt = f"Summarize the following content in 2 sentences:\n\n{combined_text}"
    response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=prompt,
        max_tokens=100,
        temperature=0.5,
    )
    summary = response.choices[0].text.strip()
    return summary

def compare_industries(source_summary: str) -> int:
    prompt = f"""
Compare the industry and purpose of the following source domain with RS Online:

Source Domain: {source_summary}

RS Online: {RS_ONLINE_DESCRIPTION}

On a scale of 0 to 100, how relevant is the source domain to RS Online in terms of industry, target audience, or potential business relationship? Provide only the numeric score.
"""
    response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=prompt,
        max_tokens=10,
        temperature=0,
    )
    try:
        return int(response.choices[0].text.strip())
    except ValueError:
        return 0

def generate_detailed_explanation(source_summary: str, industry_similarity: int) -> str:
    prompt = f"""
Source Domain Summary: {source_summary}
Industry Similarity score: {industry_similarity}

RS Online Description: {RS_ONLINE_DESCRIPTION}

Provide a brief explanation (3 sentences) of the relevancy between the source domain and RS Online based on the industry similarity score. Focus on the source domain's potential relationship with RS Online's products or services.
"""
    response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=prompt,
        max_tokens=150,
        temperature=0.7,
    )
    return response.choices[0].text.strip()

def analyse_relevancy_batch(df: pd.DataFrame, progress_bar) -> pd.DataFrame:
    column_mapping = {
        'Market': ['market', 'Market', 'MARKET'],
        'Source Domain': ['source domain', 'Source Domain', 'SOURCE DOMAIN', 'source_domain'],
        'Target URL': ['target url', 'Target URL', 'TARGET URL', 'target_url']
    }
    actual_columns = {expected_col: next((col for col in possible_names if col in df.columns), None) for expected_col, possible_names in column_mapping.items()}
    missing_columns = [col for col, found in actual_columns.items() if not found]
    if missing_columns:
        raise ValueError(f"DataFrame is missing the following required columns: {', '.join(missing_columns)}")
    df = df.rename(columns={v: k for k, v in actual_columns.items()})

    results = []
    total_rows = len(df)
    for index, row in df.iterrows():
        search_results = fetch_domain_search_results(row['Source Domain'])
        source_summary = generate_domain_explanation_from_results(search_results)
        industry_similarity = compare_industries(source_summary)
        explanation = generate_detailed_explanation(source_summary, industry_similarity)

        # Determine initial verdict based on industry similarity
        if industry_similarity < NOT_ACCEPTABLE:
            initial_verdict = "Not Acceptable"
        elif industry_similarity >= HIGH_RELEVANCY:
            initial_verdict = "High Quality Link"
        else:
            initial_verdict = "Medium Relevancy"

        results.append({
            'Market': row['Market'],
            'Source Domain': row['Source Domain'],
            'Target URL': row['Target URL'],
            'Industry Similarity': industry_similarity,
            'Explanation': explanation,
            'Our Verdict': initial_verdict
        })
        progress_bar.progress((index + 1) / total_rows)
    return pd.DataFrame(results)

def generate_excel(results_df):
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Market', 'Source Domain', 'Target URL', 'Industry Similarity', 'Explanation', 'Our Verdict']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    for row_num, data in enumerate(results_df.itertuples(index=False), start=2):
        for col_num, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            if col_num == 4:  # Industry Similarity column
                try:
                    score = int(value)
                    if score < NOT_ACCEPTABLE:
                        cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                    elif score >= HIGH_RELEVANCY:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                except ValueError:
                    pass
    workbook.save(output)
    return output.getvalue()

st.title("Bulk Backlink Relevancy Analyser")

# Main content area
uploaded_file = st.file_uploader("Upload CSV or Excel file", type=['csv', 'xlsx', 'xls'])

if uploaded_file is not None:
    try:
        if 'df' not in state:
            state.df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)

        st.subheader("Uploaded Data")
        st.dataframe(state.df, height=200)

        if st.button("Analyse Relevancy", key="analyse_button"):
            progress_bar = st.progress(0)
            with st.spinner("Analysing links..."):
                state.results_df = analyse_relevancy_batch(state.df, progress_bar)
            st.success("Analysis complete!")

        if 'results_df' in state:
            st.subheader("Analysis Results")
            for index, row in state.results_df.iterrows():
                with st.expander(f"Source: {row['Source Domain']} → Target: {row['Target URL']}", expanded=True):
                    col1, col2 = st.columns([1, 3])
                    with col1:
                        industry_score = row['Industry Similarity']
                        st.metric("Industry Similarity", industry_score)
                        if industry_score < NOT_ACCEPTABLE:
                            st.error("Not Acceptable")
                            default_verdict = "Not Acceptable"
                        elif industry_score >= HIGH_RELEVANCY:
                            st.success("High Relevancy")
                            default_verdict = "High Quality Link"
                        else:
                            st.info("Medium Relevancy")
                            default_verdict = "Medium Relevancy"

                        verdict_key = f"verdict_{index}"
                        if verdict_key not in state:
                            state[verdict_key] = row.get('Our Verdict', default_verdict)

                        verdict = st.radio(
                            "Our Verdict:",
                            ["High Quality Link", "Medium Relevancy", "Not Acceptable"],
                            key=verdict_key,
                            index=["High Quality Link", "Medium Relevancy", "Not Acceptable"].index(state[verdict_key]),
                        )

                        if state[verdict_key] != verdict:
                            state[verdict_key] = verdict
                            state.results_df.at[index, 'Our Verdict'] = verdict

                    with col2:
                        st.write("**Explanation:**")
                        st.write(row['Explanation'])

            if st.button("Update Results"):
                st.success("Results updated successfully!")

            excel_data = generate_excel(state.results_df)
            st.download_button(
                label="Download results as Excel",
                data=excel_data,
                file_name="backlink_relevancy_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except ValueError as ve:
        st.error(f"An error occurred: {str(ve)}")
        st.error("Please make sure your file contains all the required columns.")
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        st.error("Please check your file format and try again.")
else:
    st.info("Please upload a CSV or Excel file to begin the analysis.")
