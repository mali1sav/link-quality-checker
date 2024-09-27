import streamlit as st
import pandas as pd
import openai
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from streamlit import session_state as state
import os
import requests
import json
from typing import List, Dict
from dotenv import load_dotenv
import io

# Load environment variables from .env file in the current directory
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '.env'))

# Initialize OpenAI client
openai.api_key = os.getenv("OPENAI_API_KEY")
DATAFORSEO_API_KEY = os.getenv("DATAFORSEO_API_KEY")

# Check for API credentials
if not DATAFORSEO_API_KEY:
    st.error("DataForSEO API key is missing. Please check your .env file and ensure DATAFORSEO_API_KEY is set.")
    st.stop()

st.set_page_config(layout="wide", page_title="Content Relevancy Analyser")

# Constants
QUALITY_THRESHOLD = 60  # Threshold for Good Quality verdict
BATCH_SIZE = 3  # Number of domains to process in each batch

# RS Online description
RS_ONLINE_DESCRIPTION = """
RS Online, or RS Components, is a global distributor of industrial and electronic products, serving a broad range of sectors and industries with a comprehensive selection of tools, equipment, and components.

Key Target Sectors:
Manufacturing and Industrial: Supports companies in automotive, aerospace, electronics, and general manufacturing sectors, providing components for automation, control, and maintenance systems.
Engineering and R&D: Targets engineers and researchers in electronics, mechanical engineering, and product design, offering prototyping tools, test & measurement equipment, and custom automation solutions.
Healthcare and Pharmaceuticals: Supplies equipment for medical devices, diagnostic tools, and pharmaceutical manufacturing, ensuring high standards of safety and automation.
Energy and Utilities: Provides solutions for renewable energy, electrical distribution, and utility maintenance, focusing on electrification, cabling, and automation for optimal performance and safety.
Construction and Facilities Management: Supports infrastructure projects, offering products for electrical installations, HVAC, facility maintenance, and essential safety gear.
"""

# Market to location and language code mapping
MARKET_CODES = {
    'FR': {"location_code": 2250, "language_code": "fr"},
    'UK': {"location_code": 2826, "language_code": "en"},
    'IT': {"location_code": 2380, "language_code": "it"},
    'DE': {"location_code": 2276, "language_code": "de"}
}

def fetch_serp_data(domain: str, market: str) -> List[Dict]:
    url = "https://api.dataforseo.com/v3/serp/google/organic/live/advanced"
    headers = {
        'Authorization': f'Basic {DATAFORSEO_API_KEY}',
        'Content-Type': 'application/json'
    }
    market_info = MARKET_CODES.get(market, MARKET_CODES['UK'])  # Default to UK if market not found
    payload = json.dumps([{
        "keyword": f"site:{domain}",
        "location_code": market_info["location_code"],
        "language_code": market_info["language_code"],
        "device": "desktop",
        "os": "windows",
        "depth": 10  # Limit to 10 results
    }])

    try:
        response = requests.post(url, headers=headers, data=payload)
        if response.status_code == 200:
            result = response.json()
            return result['tasks'][0]['result'][0]['items'][:10]  # Return only the first 10 results
        else:
            raise Exception(f"DataForSEO SERP API request failed with status {response.status_code}: {response.text}")
    except requests.RequestException as e:
        raise Exception(f"Network error when fetching SERP data: {str(e)}")

def translate_text(text: str, target_language: str = "en") -> str:
    prompt = f"Translate the following text to English: '{text}'"
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are a professional translator."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=100
    )
    return response.choices[0].message.content.strip()

def calculate_content_similarity(source_summary: str, serp_data: List[Dict], rs_description: str, market: str) -> float:
    # Translate non-English content if necessary
    if MARKET_CODES.get(market, {}).get("language_code") != "en":
        source_summary = translate_text(source_summary)
        serp_data = [
            {
                "title": translate_text(item.get("title", "")),
                "description": translate_text(item.get("description", ""))
            }
            for item in serp_data
        ]

    serp_content = "\n".join([f"{item.get('title', '')}: {item.get('description', '')}" for item in serp_data])

    # Define the prompt
    prompt = f"""
Compare the following website content to the RS Online description:

Website Content:
{source_summary}

{serp_content}

RS Online Description:
{rs_description}

On a scale of 0 to 100, how similar is the content of the website to RS Online in terms of products, services, and target industries? Provide only the numerical score.
"""

    # Call the OpenAI API
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are an AI assistant that analyzes content similarity between websites. Always respond with a number between 0 and 100."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=10
    )

    try:
        similarity_score = float(response.choices[0].message.content.strip())
        return max(0, min(100, similarity_score))  # Ensure the score is between 0 and 100
    except ValueError:
        st.warning(f"Invalid similarity score received: {response.choices[0].message.content.strip()}. Using default value of 50.")
        return 50.0


def generate_source_summary(domain: str) -> str:
    prompt = f"Provide a one-sentence summary of what the website {domain} is about, based on its domain name."
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are an AI assistant that provides concise summaries of websites based on their domain names."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=50
    )
    return response.choices[0].message.content.strip()

def generate_domain_explanation(domain: str, content_similarity: float) -> str:
    prompt = f"""
    Domain: {domain}
    Content Similarity to RS Online: {content_similarity:.2f}%

    Provide a brief explanation (3-4 sentences) about the website's content relevance to RS Online.
    """
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are an AI assistant that provides concise explanations about websites based on their content similarity."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=80
    )
    return response.choices[0].message.content.strip()

def analyse_relevancy_batch(df: pd.DataFrame, progress_bar, batch_size: int = BATCH_SIZE):
    column_mapping = {
        'Market': ['market', 'Market', 'MARKET'],
        'Source Domain': ['source domain', 'Source Domain', 'SOURCE DOMAIN', 'source_domain'],
        'Target URL': ['target url', 'Target URL', 'TARGET URL', 'target_url']
    }
    actual_columns = {expected_col: next((col for col in possible_names if col in df.columns), None) for expected_col, possible_names in column_mapping.items()}
    missing_columns = [col for col, found in actual_columns.items() if not found]
    if missing_columns:
        raise ValueError(f"DataFrame is missing the following required columns: {', '.join(missing_columns)}")
    df = df.rename(columns={v: k for k, v in actual_columns.items()})

    results = []
    total_rows = len(df)
    total_index = 0  # Initialize total_index for unique widget keys

    for start_idx in range(0, total_rows, batch_size):
        batch = df.iloc[start_idx:start_idx + batch_size]
        batch_results = []

        for _, row in batch.iterrows():
            try:
                serp_data = fetch_serp_data(row['Source Domain'], row['Market'])
                source_summary = generate_source_summary(row['Source Domain'])
                content_similarity = calculate_content_similarity(source_summary, serp_data, RS_ONLINE_DESCRIPTION, row['Market'])
                explanation = generate_domain_explanation(row['Source Domain'], content_similarity)

                verdict = "Relevant" if content_similarity >= QUALITY_THRESHOLD else "Not Relevant"

                result = {
                    'Market': row['Market'],
                    'Source Domain': row['Source Domain'],
                    'Target URL': row['Target URL'],
                    'Content Similarity': f"{content_similarity:.2f}%",
                    'Explanation': explanation,
                    'Our Verdict': verdict,
                    'Source Summary': source_summary,
                    'SERP Titles': [item.get('title', '') for item in serp_data],
                    'SERP Descriptions': [item.get('description', '') for item in serp_data]
                }
                batch_results.append(result)

            except Exception as e:
                error_message = f"Error processing {row['Source Domain']}: {str(e)}"
                st.error(error_message)
                batch_results.append({
                    'Market': row['Market'],
                    'Source Domain': row['Source Domain'],
                    'Target URL': row['Target URL'],
                    'Content Similarity': 'Error',
                    'Explanation': error_message,
                    'Our Verdict': 'Error',
                    'Source Summary': 'Error',
                    'SERP Titles': [],
                    'SERP Descriptions': []
                })

            total_index += 1  # Increment total_index
            progress_bar.progress(total_index / total_rows)

        results.extend(batch_results)
        st.success(f"Processed batch {start_idx // batch_size + 1} of {(total_rows + batch_size - 1) // batch_size}")

    # After processing all batches, update state and reset index
    state.results_df = pd.DataFrame(results).reset_index(drop=True)
    return state.results_df

def display_results(results_df):
    st.subheader("Analysis Results")
    for idx, row in results_df.iterrows():
        with st.expander(f"Source: {row['Source Domain']} → Target: {row['Target URL']}", expanded=True):
            col1, col2 = st.columns([1, 3])
            with col1:
                st.metric("Content Similarity", row['Content Similarity'])

                verdict_key = f"verdict_{idx}"
                if verdict_key not in state:
                    state[verdict_key] = row['Our Verdict']

                verdict = st.radio(
                    "Our Verdict:",
                    ["Relevant", "Not Relevant"],
                    key=f"verdict_radio_{idx}",  # Use a unique key for each radio button
                    index=["Relevant", "Not Relevant"].index(state[verdict_key]),
                    on_change=lambda: None  # Prevent collapsing on change
                )

                if state[verdict_key] != verdict:
                    state[verdict_key] = verdict
                    results_df.at[idx, 'Our Verdict'] = verdict

            with col2:
                st.write("**Explanation:**")
                st.write(row['Explanation'])
                st.write("**Source Summary:**")
                st.write(row['Source Summary'])
                st.write("**SERP Titles:**")
                st.write(", ".join(row['SERP Titles']))
                st.write("**SERP Descriptions:**")
                st.write(", ".join(row['SERP Descriptions']))

def generate_excel(results_df):
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Market', 'Source Domain', 'Target URL', 'Content Similarity', 'Explanation', 'Our Verdict', 'Source Summary', 'SERP Titles', 'SERP Descriptions']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    for row_idx, data in enumerate(results_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=str(value) if isinstance(value, list) else value)
            if col_idx == 4:  # Content Similarity column
                try:
                    score = float(value.strip('%'))
                    if score < QUALITY_THRESHOLD:
                        cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                except ValueError:
                    pass  # If value is not a float (e.g., 'Error'), do nothing
    workbook.save(output)
    return output.getvalue()

st.title("Content Relevancy Analyser")

# Main content area
uploaded_file = st.file_uploader("Upload CSV or Excel file", type=['csv', 'xlsx', 'xls'])

if uploaded_file is not None:
    try:
        if 'df' not in state:
            state.df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)

        st.subheader("Uploaded Data")
        st.dataframe(state.df, height=200)

        if st.button("Analyse Content", key="analyse_button"):
            progress_bar = st.progress(0)
            with st.spinner("Analysing content..."):
                state.results_df = analyse_relevancy_batch(state.df, progress_bar)
            st.success("Analysis complete!")
            display_results(state.results_df)

        if 'results_df' in state:
            display_results(state.results_df)

            if st.button("Update Results"):
                st.success("Results updated successfully!")

            excel_data = generate_excel(state.results_df)
            st.download_button(
                label="Download results as Excel",
                data=excel_data,
                file_name="content_relevancy_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except ValueError as ve:
        st.error(f"An error occurred: {str(ve)}")
        st.error("Please make sure your file contains all the required columns.")
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        st.error("Please check your file format and try again.")
else:
    st.info("Please upload a CSV or Excel file to begin the analysis.")
